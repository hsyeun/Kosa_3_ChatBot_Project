import pymysql
import openpyxl

from Config.DatabaseConfig import * # DB 접속 정보 불러오기


# 학습 데이터 초기화
def all_clear_train_data(db):
    # 기존 학습 데이터 삭제
    sql = '''
            delete from chatbot_qna_data
        '''
    with db.cursor() as cursor:
        cursor.execute(sql)

    # auto increment 초기화
    sql = '''
    ALTER TABLE chatbot_qna_data AUTO_INCREMENT=1
    '''
    with db.cursor() as cursor:
        cursor.execute(sql)


# db에 데이터 저장
def insert_data(db, xls_row):
    intent, ner, query, answer = xls_row

    sql = '''
        INSERT chatbot_qna_data(intent, ner, query, answer) 
        values(
         '%s', '%s', '%s', '%s'
        )
    ''' % (intent.value, ner.value, query.value, answer.value)

    # 엑셀에서 불러온 cell에 데이터가 없는 경우, null 로 치환
    sql = sql.replace("'None'", "null")

    with db.cursor() as cursor:
        cursor.execute(sql)
        # print('{} 저장'.format(query.value))
        db.commit()


train_file = r'C:\dev\Project_chatbot\Kosa_3_ChatBot_Project\train_tools\qna\DB_QnA.xlsx'
db = None
try:
    db = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        passwd=DB_PASSWORD,
        db=DB_NAME,
        charset='utf8'
    )

    # 기존 학습 데이터 초기화
    all_clear_train_data(db)

    # df = pd.read_excel(train_file)
    # print(df.info())

    # 학습 엑셀 파일 불러오기
    wb = openpyxl.load_workbook(train_file)
    ws = wb.sheetnames
    print(ws)
    sheet = wb['Sheet1']
    for row in sheet.iter_rows(min_row=2, min_col=2): # 해더는 불러오지 않음
        # 데이터 저장
        insert_data(db, row)

    wb.close()

except Exception as e:
    print(e)

finally:
    if db is not None:
        db.close()

